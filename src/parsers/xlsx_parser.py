"""
XLSX Document Parser - FIXED: Path Traversal Protection
Extracts text and formatting from Excel .xlsx files.
"""
from pathlib import Path
from typing import List, Dict
import logging
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

from ..core.interfaces import IDocumentParser
from ..core.models import (
    Document,
    TextSegment,
    FileType,
    SegmentType,
    SegmentPosition,
    CellFormatting,
    DocumentMetadata,
    DocumentStyles
)


logger = logging.getLogger(__name__)


class XlsxParser(IDocumentParser):
    """Parser for Microsoft Excel .xlsx files."""
    
    def __init__(self, workspace_root: Path = None):
        """
        Initialize parser.
        
        Args:
            workspace_root: Root directory for allowed file operations.
                           If None, uses current working directory.
        """
        self.workspace_root = (workspace_root or Path.cwd()).resolve()
    
    @property
    def supported_file_type(self) -> FileType:
        """Supported file type."""
        return FileType.XLSX
    
    def can_parse(self, file_path: Path) -> bool:
        """Check if file can be parsed."""
        return file_path.suffix.lower() == '.xlsx'
    
    def validate_document(self, file_path: Path) -> bool:
        """Validate document before parsing."""
        try:
            # SECURITY FIX: Validate path to prevent path traversal
            file_path = self._validate_and_resolve_path(file_path)
        except ParsingError as e:
            logger.error(f"Path validation failed: {e}")
            return False
        
        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            return False
        
        if not self.can_parse(file_path):
            logger.error(f"Unsupported file type: {file_path.suffix}")
            return False
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            return True
        except Exception as e:
            logger.error(f"Cannot open workbook: {e}")
            return False
    
    def parse(self, file_path: Path) -> Document:
        """
        Parse XLSX document.
        
        Args:
            file_path: Path to .xlsx file
            
        Returns:
            Document with extracted content and formatting
        """
        # SECURITY FIX: Validate path first
        file_path = self._validate_and_resolve_path(file_path)
        
        if not self.validate_document(file_path):
            raise ParsingError(f"Invalid document: {file_path}")
        
        wb = None
        try:
            # Load workbook (not read-only to access formatting)
            wb = load_workbook(file_path, data_only=False)
            
            # Extract metadata
            metadata = self._extract_metadata(wb)
            
            # Extract styles
            styles = self._extract_styles(wb)
            
            # Extract segments from all sheets
            segments = []
            
            for sheet in wb.worksheets:
                sheet_segments = self._extract_sheet_segments(sheet)
                segments.extend(sheet_segments)
            
            document = Document(
                file_path=file_path,
                file_type=FileType.XLSX,
                segments=segments,
                metadata=metadata,
                styles=styles
            )
            
            logger.info(
                f"Parsed {file_path.name}: "
                f"{len(wb.worksheets)} sheets, "
                f"{len(segments)} segments, "
                f"{document.total_words} words"
            )
            
            return document
            
        except Exception as e:
            logger.error(f"Failed to parse document: {e}")
            raise ParsingError(f"Parsing failed: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
    
    def extract_segments(self, document: Document) -> List[TextSegment]:
        """Extract translatable segments."""
        # Filter out empty segments and numeric-only cells
        segments = [
            s for s in document.segments 
            if s.text.strip() and not self._is_numeric_only(s.text)
        ]
        
        logger.info(f"Extracted {len(segments)} translatable segments")
        return segments
    
    def _validate_and_resolve_path(self, file_path: Path) -> Path:
        """
        SECURITY: Validate and resolve file path to prevent path traversal.
        
        Args:
            file_path: Input path
            
        Returns:
            Resolved absolute path
            
        Raises:
            ParsingError: If path is unsafe
        """
        # Resolve to absolute path
        file_path = file_path.resolve()
        
        # Check workspace boundary
        try:
            file_path.relative_to(self.workspace_root)
        except ValueError:
            raise ParsingError(
                f"Access denied: Path outside workspace.\n"
                f"File: {file_path}\n"
                f"Workspace: {self.workspace_root}"
            )
        
        # Validate filename
        self._validate_filename(file_path.name)
        
        return file_path
    
    def _validate_filename(self, filename: str) -> None:
        """
        Validate filename for cross-platform compatibility.
        
        Args:
            filename: Filename to validate
            
        Raises:
            ParsingError: If filename is invalid
        """
        # OS-specific dangerous characters
        if os.name == 'nt':  # Windows
            dangerous = ['<', '>', ':', '"', '/', '\\', '|', '?', '*', '\0']
        else:
            dangerous = ['/', '\0']
        
        for char in dangerous:
            if char in filename:
                raise ParsingError(f"Invalid character in filename: {repr(char)}")
        
        # Windows reserved names
        if os.name == 'nt':
            reserved = {
                'CON', 'PRN', 'AUX', 'NUL',
                'COM1', 'COM2', 'COM3', 'COM4', 'COM5', 'COM6', 'COM7', 'COM8', 'COM9',
                'LPT1', 'LPT2', 'LPT3', 'LPT4', 'LPT5', 'LPT6', 'LPT7', 'LPT8', 'LPT9'
            }
            name_without_ext = filename.rsplit('.', 1)[0].upper()
            if name_without_ext in reserved:
                raise ParsingError(f"Reserved Windows filename: {filename}")
            
            # No trailing spaces or dots
            if filename.rstrip() != filename or filename.rstrip('.') != filename:
                raise ParsingError("Filename cannot end with space or dot on Windows")
        
        # Length check
        if len(filename) > 255:
            raise ParsingError("Filename too long (max 255 characters)")
    
    def _extract_metadata(self, wb) -> DocumentMetadata:
        """Extract workbook metadata."""
        props = wb.properties
        
        return DocumentMetadata(
            title=props.title,
            author=props.creator,
            subject=props.subject,
            keywords=props.keywords,
            comments=props.description,
            created=props.created,
            modified=props.modified,
            last_modified_by=props.lastModifiedBy
        )
    
    def _extract_styles(self, wb) -> DocumentStyles:
        """Extract workbook-level styles."""
        styles = DocumentStyles()
        
        # Extract default font from first cell
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=1, max_col=1):
                for cell in row:
                    if cell.font:
                        styles.default_font = cell.font.name
                        return styles
        
        return styles
    
    def _extract_sheet_segments(self, sheet) -> List[TextSegment]:
        """Extract segments from a single sheet."""
        segments = []
        sheet_name = sheet.title
        
        for row in sheet.iter_rows():
            for cell in row:
                # Skip empty cells
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                
                # Skip formulas (we want calculated values)
                text = str(cell.value)
                
                # Create segment
                segment = TextSegment(
                    id=f"sheet_{sheet_name}_row_{cell.row}_col_{cell.column}",
                    text=text,
                    segment_type=SegmentType.TABLE_CELL,
                    position=SegmentPosition(
                        sheet_name=sheet_name,
                        row=cell.row,
                        column=cell.column
                    ),
                    cell_formatting=self._extract_cell_formatting(cell, sheet)
                )
                segments.append(segment)
        
        return segments
    
    def _extract_cell_formatting(self, cell, sheet) -> CellFormatting:
        """Extract formatting from cell."""
        font = cell.font
        fill = cell.fill
        alignment = cell.alignment
        border = cell.border
        
        # Font color
        font_color = None
        if font and font.color and hasattr(font.color, 'rgb'):
            rgb = font.color.rgb
            if rgb and len(rgb) >= 6:
                # Remove alpha channel if present (ARGB -> RGB)
                if len(rgb) == 8:
                    rgb = rgb[2:]
                font_color = f"#{rgb}"
        
        # Fill color
        fill_color = None
        if fill and fill.start_color and hasattr(fill.start_color, 'rgb'):
            rgb = fill.start_color.rgb
            if rgb and len(rgb) >= 6:
                if len(rgb) == 8:
                    rgb = rgb[2:]
                fill_color = f"#{rgb}"
        
        # Alignment
        h_align = None
        v_align = None
        wrap = False
        if alignment:
            h_align = alignment.horizontal
            v_align = alignment.vertical
            wrap = alignment.wrap_text or False
        
        # Border
        border_style = None
        if border:
            border_style = {
                'top': border.top.style if border.top else None,
                'bottom': border.bottom.style if border.bottom else None,
                'left': border.left.style if border.left else None,
                'right': border.right.style if border.right else None
            }
        
        # Column width and row height
        col_letter = cell.column_letter
        column_width = sheet.column_dimensions[col_letter].width if col_letter in
        column_width = sheet.column_dimensions[col_letter].width if col_letter in sheet.column_dimensions else None
        row_height = sheet.row_dimensions[cell.row].height if cell.row in sheet.row_dimensions else None
        
        return CellFormatting(
            font_name=font.name if font else None,
            font_size=font.size if font else None,
            font_bold=font.bold if font else False,
            font_italic=font.italic if font else False,
            font_color=font_color,
            fill_color=fill_color,
            fill_pattern=fill.patternType if fill else None,
            horizontal_alignment=h_align,
            vertical_alignment=v_align,
            wrap_text=wrap,
            border_style=border_style,
            number_format=cell.number_format if cell.number_format != 'General' else None,
            column_width=column_width,
            row_height=row_height
        )
    
    def _is_numeric_only(self, text: str) -> bool:
        """Check if text is numeric only (don't translate numbers)."""
        try:
            # Try to convert to number
            float(text.replace(',', '').replace(' ', ''))
            return True
        except:
            return False


class ParsingError(Exception):
    """Exception raised when parsing fails."""
    pass
